from typing import Dict, Any, List
from threading import RLock
import openpyxl
from src.utils.logger import logger
from pathlib import Path


class ExcelUpdater:
    """
    A thread-safe utility class for applying updates to Excel workbooks.

    This class handles all Excel update operations in a centralized way, supporting different
    types of updates including:
    1. Product name based updates
    2. Row/column based updates
    3. Date based updates in specific sheets

    All operations are protected by a global lock to ensure thread safety.
    """

    _global_lock = RLock()  # Class-level lock for all ExcelUpdater instances

    def __init__(self, output_path: Path):
        """
        Initialize the Excel updater.

        Args:
            output_path: Path where the workbook should be saved

        Raises:
            FileNotFoundError: If the workbook doesn't exist
            openpyxl.utils.exceptions.InvalidFileException: If the file is not a valid Excel file
        """
        self.workbook = openpyxl.load_workbook(output_path)
        self.output_path = output_path

    def open_workbook(self, output_path: Path) -> None:
        """
        Open the workbook from the specified path.
        """
        self.workbook = openpyxl.load_workbook(output_path)
        self.output_path = output_path

    def save_workbook(self) -> None:
        """
        Save the workbook to the specified output path.
        Thread-safe operation using the global lock.

        Raises:
            PermissionError: If the file is locked or permission denied
            Exception: For other IO related errors
        """
        try:
            with ExcelUpdater._global_lock:
                self.workbook.save(self.output_path)
                logger.info(
                    f"Successfully saved workbook to: {self.output_path}")
        except Exception as e:
            logger.error(f"Error saving workbook: {str(e)}")
            raise

    def close_workbook(self) -> None:
        """
        Close the workbook.
        Thread-safe operation using the global lock.

        Raises:
            Exception: If there's an error closing the workbook
        """
        try:
            with ExcelUpdater._global_lock:
                self.workbook.close()
                logger.info("Successfully closed workbook")
        except Exception as e:
            logger.error(f"Error closing workbook: {str(e)}")
            raise

    def _validate_update(self, update: Dict[str, Any]) -> None:
        """
        Validate a single update instruction.

        Args:
            update: The update instruction to validate

        Raises:
            ValueError: If the update instruction is invalid
        """
        if 'sheet' not in update:
            raise ValueError("Update must specify a sheet name")

        if update['sheet'] not in self.workbook.sheetnames:
            raise ValueError(
                f"Sheet '{update['sheet']}' not found in workbook")

        if update['sheet'] in ('调价前', '调价后'):
            if 'product_name' in update:
                if not all(key in update for key in ['column', 'value']):
                    raise ValueError(
                        "Product name based update must include column and value")
            elif 'updates' in update:
                if not all(all(key in u for key in ['row', 'column', 'value']) for u in update['updates']):
                    raise ValueError(
                        "Row/column based updates must include row, column, and value")
            else:
                raise ValueError("Invalid update format for sheet '调价前'")

        elif update['sheet'] == '油品优惠明细 2':
            if 'date' not in update or 'updates' not in update:
                raise ValueError(
                    "Date based update must include date and updates")
            if not all(all(key in u for key in ['column', 'value']) for u in update['updates']):
                raise ValueError(
                    "Date based updates must include column and value")

    def apply_updates(self, updates: List[Dict[str, Any]]) -> None:
        """
        Apply updates to the workbook in a thread-safe manner.

        Args:
            updates: List of update instructions. Each update can have the following formats:
            1. {'sheet': '调价前', 'product_name': str, 'column': str, 'value': float} - Product name based update
            2. {'sheet': '调价前', 'updates': [{'row': int, 'column': str, 'value': float}]} - Row/column based update
            3. {'sheet': '油品优惠明细 2', 'date': int, 'updates': [{'column': str, 'value': float}]} - Date based update

        Raises:
            ValueError: If update format is invalid
            Exception: For other update related errors
        """
        try:
            with ExcelUpdater._global_lock:
                # Initialize handling_fees dictionary at the start
                handling_fees = {}

                for update in updates:
                    # Validate update format
                    self._validate_update(update)

                    sheet_name = update['sheet']
                    sheet = self.workbook[sheet_name]

                    if sheet_name in ('调价前', '调价后'):
                        if 'section' in update:
                            target_section = update['section']
                            if target_section in ['A', 'B', 'C']:
                                current_section = None
                                # Handle product name based updates
                                for row in sheet.iter_rows(min_row=3):
                                    if row[0].value is not None:
                                        current_section = row[0].value.strip()
                                    if current_section == target_section:
                                        # B column data
                                        product_name = row[1].value
                                        if not product_name:  # Skip empty rows
                                            continue
                                        # Normalize both product names before comparison
                                        normalized_row_name = self._normalize_product_name(
                                            product_name)
                                        normalized_update_name = self._normalize_product_name(
                                            update['product_name'])
                                        if normalized_row_name == normalized_update_name:
                                            col_idx = self._get_column_index(
                                                update['column'])
                                            row[col_idx].value = update['value']
                                            break

                        elif 'updates' in update:
                            # Handle row/column based updates
                            for row_update in update['updates']:
                                row_idx = row_update['row'] - 1
                                col_idx = self._get_column_index(
                                    row_update['column'])

                                # Special handling for handling fees (row 81, column E)
                                if row_idx + 1 == 81 and row_update['column'] == 'E':
                                    # Track handling fees for accumulation
                                    cell_key = (row_idx + 1, col_idx + 1)
                                    if cell_key not in handling_fees:
                                        handling_fees[cell_key] = 0
                                        # Clear existing value when first encountering this cell
                                        sheet.cell(
                                            row=row_idx + 1, column=col_idx + 1).value = None

                                    # Accumulate the new value
                                    handling_fees[cell_key] += row_update['value']
                                    # Update the cell with accumulated value
                                    sheet.cell(
                                        row=row_idx + 1, column=col_idx + 1).value = handling_fees[cell_key]
                                else:
                                    # For all other cells, just set the new value
                                    cell = sheet.cell(
                                        row=row_idx + 1, column=col_idx + 1)
                                    cell.value = row_update['value']

                    elif sheet_name == '油品优惠明细 2':
                        # Handle date based updates
                        date = update.get('date')
                        date_found = False

                        # Find the row corresponding to the date
                        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                            if row[1].value == date:  # B column is date
                                date_found = True
                                for col_update in update['updates']:
                                    col_idx = self._get_column_index(
                                        col_update['column'])
                                    row[col_idx].value = col_update['value']
                                break

                        if not date_found:
                            logger.warning(
                                f"Date {date} not found in sheet {sheet_name}")

                logger.info("Successfully applied updates to workbook")

        except Exception as e:
            logger.error(f"Error applying updates: {str(e)}")
            raise

    def _get_column_index(self, column: str) -> int:
        """
        Convert Excel column letter(s) to zero-based index.
        Supports both single letter (A-Z) and double letter (AA-ZZ) columns.

        Args:
            column: Column letter(s) (e.g., 'A', 'B', 'AA', 'AB', etc.)

        Returns:
            Zero-based column index
        """
        result = 0
        for char in column.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1  # Convert to 0-based index

    @staticmethod
    def _normalize_product_name(name: str) -> str:
        """
        Normalize product name (油枪序号) to a standard format.
        Removes '号' suffix if present and returns the number as string.

        Args:
            name: The product name to normalize

        Returns:
            Normalized product name
        """
        if not name:
            return ""
        # Convert to string and remove any whitespace
        name_str = str(name).strip()
        # Remove '号' suffix if present
        if name_str.endswith('号'):
            name_str = name_str[:-1]
        return name_str
